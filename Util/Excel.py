#encoding =utf-8
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import traceback
import time

class ParseExcel(object):
    """"此类主要封装常用的操作excel文件的方法"""

    def __init__(self,excel_file_path):
        if not os.path.exists(excel_file_path):
            self.wb = None
        self.excel_file_path = excel_file_path
        self.wb =  load_workbook(excel_file_path)
        self.sheet = self.set_sheet_by_name(self.wb.sheetnames[0])

    def get_excel_file_path(self):
        return self.excel_file_path

    def set_sheet_by_name(self,name):
        if name in self.wb.sheetnames:
            self.sheet  = self.wb[name]
            return self.sheet
        self.sheet = None
        return self.sheet

    def get_sheet_by_name(self,name):
        return self.wb[name]

    def get_all_sheet_names(self):
        return self.wb.sheetnames

    def get_all_sheet_objects(self):
        sheets=[]
        for sheetname in self.wb.sheetnames:
            sheets.append(self.get_sheet_by_name(sheetname))
        return sheets

    def get_current_sheet_name(self):
        if self.sheet is not None:
            return self.sheet.title
        return None

    def get_min_row(self):#从1开始计数
        try:
            return self.sheet.min_row
        except:
            traceback.print_exc()
            return None

    def get_max_row(self):
        try:
            return self.sheet.max_row
        except:
            traceback.print_exc()
            return None

    def get_min_col(self):#从1开始计数
        try:
            return self.sheet.min_column
        except:
            traceback.print_exc()
            return None

    def get_max_col(self):
        try:
            return self.sheet.max_column
        except:
            traceback.print_exc()
            return None

    def get_row(self,row_no):
        if not isinstance(row_no,int):
            return None
        try:
            return list(self.sheet.rows)[row_no - 1]
        except:
            traceback.print_exc()


    def get_col(self,col_no):
        if not isinstance(col_no,int):
            return None
        try:
            return list(self.sheet.columns)[col_no - 1]
        except:
            traceback.print_exc()

    def get_cell_value(self,row_no,col_no):
        """参数行号和列表从1开始表示第一行"""
        if (not isinstance(row_no,int)) or (not isinstance(col_no,int)):
            return None

        try:
            return self.sheet.cell(row=row_no, column=col_no).value
        except:
            traceback.print_exc()

    def write_cell(self,row_no,col_no,content):
        """参数行号和列表从1开始表示第一行"""
        if (not isinstance(row_no, int)) or (not isinstance(col_no, int)):
            return None
        try:
            self.sheet.cell(row=row_no, column=col_no).value=content
            self.wb.save(self.excel_file_path)
        except:
            traceback.print_exc()

    def write_cell_date(self, row_no, col_no):
        timeTup = time.localtime()
        currentDate = str(timeTup.tm_year) + "年" + \
                      str(timeTup.tm_mon) + "月" + str(timeTup.tm_mday)+"日"
        self.write_cell(row_no, col_no,currentDate)


    def write_cell_time(self, row_no, col_no):
        timeTup = time.localtime()
        currentTime = str(timeTup.tm_hour) + "时" + \
                      str(timeTup.tm_min) + "分" + str(timeTup.tm_sec) + "秒"
        self.write_cell(row_no, col_no, currentTime)

    def write_cell_datetime(self, row_no, col_no):
        timeTup = time.localtime()
        currentDate = str(timeTup.tm_year) + "年" + \
                      str(timeTup.tm_mon) + "月" + str(timeTup.tm_mday) + "日"
        currentTime = str(timeTup.tm_hour) + "时" + \
                      str(timeTup.tm_min) + "分" + str(timeTup.tm_sec) + "秒"
        self.write_cell(row_no, col_no, currentDate+" "+currentTime)

if __name__ == "__main__":
    #wb = ParseExcel("e:\\126邮箱联系人.xlsx")
    wb = ParseExcel(r"E:\keyword_driven_proj\TestData\测试用例.xlsx")
    #print(wb.get_col(col_no =1))
    # print(wb.get_min_row())
    # print(wb.get_max_row())
    # print(wb.get_min_col())
    # print(wb.get_max_col())
    # print(wb.get_row(1))
    # print(wb.get_col(1))
    # print(wb.get_cell_value(1,1))
    #print(wb.get_cell_value("联系人",10,4))
    #wb.write_cell(1,1,"光荣之路")
    # wb.write_cell_date(1,1)
    # wb.write_cell_time(1, 2)
    # wb.write_cell_datetime(1,3)
    print(wb.get_sheet_by_name("测试用例"))
    print(wb.get_all_sheet_objects())
    print(wb.get_all_sheet_names())